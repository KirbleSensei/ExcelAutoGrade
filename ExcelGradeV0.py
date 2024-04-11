import os

import openpyxl
import patoolib
from os import listdir
from os.path import join


# Made by Emre KAPLAN


def assertEqualsCell(pathToFolder, SheetName, Cell, expectedValue):
    """ Asserts that a cell is equal to the expected """
    for filename in listdir(pathToFolder):
        full_path = join(pathToFolder, filename)
        if os.path.isfile(full_path) and filename.endswith(".xlsx"):
            wb = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
            counter = 0
            ws = wb[SheetName]  # Open up Sheet
            if ws[Cell].value == expectedValue:
                return True


def get_cells_in_range(pathToExcel, SheetName, CellRange):
    cell_container = []
    wb = openpyxl.load_workbook(pathToExcel, read_only=True, data_only=True)
    ws = wb[SheetName]
    target_cells = ws[CellRange]
    for value_row in target_cells:
        for cell in value_row:
            cell_container.append(cell)
    return cell_container


def extract_nested_archives(path):
    """Extracts nested archives."""
    first_extract = patoolib.extract_archive(path)
    for filename in listdir(first_extract):
        full_path = join(first_extract, filename)
        patoolib.extract_archive(full_path)


def assertEqualsCells(pathToZip, SheetName, CellRange, expectedValues, WhitelistedFormulas):
    """ Asserts that a range of cells is equal to the expected tuple
    :param pathToZip: Raw path to the Initial Zip file
    :param SheetName: Name of the Sheet to read the data from
    :param CellRange: Range of Cells to read
    :param expectedValues: Tuple of expected return values of the formulas
    :param WhitelistedFormulas: Tuple of expected formulas
    """
    valueTestPassed = False
    current_directory = os.path.dirname(os.path.abspath(__file__))
    extract_nested_archives(pathToZip)

    warning_file = open("Warnings.txt", "w")  # Initialize Grading and Warning files
    grades_file = open("Grades.txt", "w")

    for filename in listdir(current_directory):  # Loop through files
        full_path = join(current_directory, filename)

        # Check if the file is an Excel file
        if os.path.isfile(full_path) and filename.endswith(".xlsx"):
            student_number = filename.split(".")[0]

            # First VALUE check
            fetched_cells = get_cells_in_range(full_path, SheetName, CellRange)
            grade = 0       # Initialize grade
            for cell in fetched_cells:
                if cell.value in expectedValues:
                    valueTestPassed = True

            # Second FORMULA check
            wb = openpyxl.load_workbook(full_path, read_only=True, data_only=False)
            ws = wb[SheetName]  # Open up Sheet
            target_cells = ws[CellRange]  # Cells to be read

            for row in target_cells:
                for cell in row:
                    if valueTestPassed:
                        if cell.value in WhitelistedFormulas:
                            grade += 1
                        else:
                            if type(cell.value) is not int and cell.value not in WhitelistedFormulas:
                                warning_file.write(
                                    "WARNING: Expected Value of student {0} is correct but used formula is not in the "
                                    "whitelist | Cell: {1} | Formula Used: {2}\n".format(
                                        student_number, cell.coordinate, cell.value))
                    else:
                        grade -= 1

            grades_file.write("Student {0}'s grade is {1}\n".format(student_number, grade))


path = r"C:\Users\Emre K\Documents\GitHub\ExcelAutoGrade\Project01.rar"  # Path to the FOLDER that contains excel files

expected = (46, 47, 197)  # List of expected values MUST BE IN THE SAME ORDER AS CELLS

whitelist = ("=SUM(D2:D12)", "=SUM(E2:E12)", "=SUM(F2:F12)")

assertEqualsCells(path, "Sheet1", "D13:F13", expected, whitelist)
