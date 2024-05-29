import os
import openpyxl
import patoolib
from os import listdir
from os.path import join
import tempfile


def assert_equals_cell(path_to_folder, sheet_name, cell, expected_value):
    """Asserts that a cell is equal to the expected value"""
    for filename in listdir(path_to_folder):
        full_path = join(path_to_folder, filename)
        if os.path.isfile(full_path) and filename.endswith(".xlsx"):
            # Load Excel file with read-only mode and data-only option
            wb = openpyxl.load_workbook(full_path, read_only=True, data_only=True)
            ws = wb[sheet_name]  # Open up Sheet
            if ws[cell].value == expected_value:
                return True


def get_cells_in_range(path_to_excel, sheet_name, cell_range):
    """Returns a list of cell objects with values from the specified cell range"""
    cell_container = []
    wb = openpyxl.load_workbook(path_to_excel, read_only=True, data_only=True)
    ws = wb[sheet_name]
    target_cells = ws[cell_range]
    for value_row in target_cells:
        for cell in value_row:
            cell_container.append(cell)
    return cell_container


def get_formulas_in_range(path_to_excel, sheet_name, cell_range):
    """Returns a list of cell objects with formulas from the specified cell range"""
    cell_container = []
    wb = openpyxl.load_workbook(path_to_excel, read_only=True, data_only=False)
    ws = wb[sheet_name]
    target_cells = ws[cell_range]
    for value_row in target_cells:
        for cell in value_row:
            cell_container.append(cell)
    return cell_container


def assert_equals_cells(path_to_zip, sheet_name, cell_range, expected_values, whitelisted_formulas):
    """
    Asserts that a range of cells is equal to the expected tuple.

    :param path_to_zip: Raw path to the Initial Zip file
    :param sheet_name: Name of the Sheet to read the data from
    :param cell_range: Range of Cells to read
    :param expected_values: Tuple of expected return values of the formulas
    :param whitelisted_formulas: Tuple of expected formulas
    """
    graded_files = []
    # Flag to track if the value test passed
    value_test_passed = False
    # Save the current directory
    original_directory = os.getcwd()

    # Use the temp directory and assign it to current_directory
    with tempfile.TemporaryDirectory() as tmpdirname:
        current_directory = tmpdirname
        # os.chdir changes the root directory of the code to the current_directory which is temp
        os.chdir(current_directory)
        # Extraction of the Project01.rar, first_extract is the path of the extracted file
        first_extract = patoolib.extract_archive(path_to_zip, verbosity=-1)
        for filename in listdir(first_extract):
            full_path = join(first_extract, filename)
            patoolib.extract_archive(full_path, verbosity=-1)

        # Open files for writing warnings and grades
        with open("Warnings.txt", "w") as warning_file:
            # Loop through files in the current directory
            for filename in listdir(current_directory):
                full_path = join(current_directory, filename)
                # Check if the file is an Excel file
                if os.path.isfile(full_path) and filename.endswith(".xlsx"):
                    student_number = filename.split(".")[0]
                    student_file = filename
                    # Get cell values within the specified range
                    fetched_value_cells = get_cells_in_range(full_path, sheet_name, cell_range)
                    grade = 0
                    # Check if fetched cell values are in the expected values
                    for cell in fetched_value_cells:
                        if cell.value in expected_values:
                            value_test_passed = True

                    # Get cell formulas within the specified range
                    fetched_formula_cells = get_formulas_in_range(full_path, sheet_name, cell_range)
                    # Loop through fetched formula cells
                    for cell in fetched_formula_cells:
                        # If the value test passed
                        if value_test_passed:
                            # Check if the formula is in the whitelist
                            if cell.value in whitelisted_formulas:
                                grade += 1
                            else:
                                # Write a warning if the formula is not in the whitelist
                                if type(cell.value) is not int and cell.value not in whitelisted_formulas:
                                    warning_file.write(
                                        "WARNING: Expected Value of student {0} is correct but used formula is not in the "
                                        "whitelist | Cell: {1} | Formula Used: {2}\n".format(
                                            student_number, cell.coordinate, cell.value))
                        else:
                            # Decrease grade if the value test didn't pass
                            grade -= 1

                    # Write the student's grade to the grades file
                    with open("Grade.txt", "w") as grades_file:
                        grades_file.write("Student {0}'s grade is {1}\n".format(student_number, grade))
                        grades_file.close()
                        patoolib.create_archive("{0} Graded.rar".format(student_number), (grades_file.name, student_file))
                        os.remove(grades_file.name)
        for filename in listdir(current_directory):
            full_path = join(current_directory, filename)
            # Check if the file is an Excel file
            if os.path.isfile(full_path) and filename.find("Graded") != -1:
                graded_files.append(filename)
        graded_files_tuple = tuple(graded_files)
        patoolib.create_archive(os.path.join(original_directory, "Graded.rar"), graded_files_tuple + ("Warnings.txt",))


# Example usage
path = r"C:\Users\Emre K\Documents\GitHub\ExcelAutoGrade\Project01.rar"

expected = (46, 47, 197)
whitelist = ("=SUM(D2:D12)", "=SUM(E2:E12)", "=SUM(F2:F12)")

assert_equals_cells(path, "Sheet1", "D13:F13", expected, whitelist)
